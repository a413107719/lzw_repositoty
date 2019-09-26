# coding=utf-8
import re
from prettytable import PrettyTable
from openpyxl.styles import Color, Fill, Font, Alignment, PatternFill, Border, Side
import openpyxl
import pandas as pd
import os
import shutil
from shutil import copy2


def get_all_excels():
    # 创建输出路径
    if os.path.isdir(outputfolder):
        shutil.rmtree(outputfolder)
        os.makedirs(outputfolder)
        print('输出文件夹已存在，已覆盖')
        print()
    else:
        os.makedirs(outputfolder)
        print('新建输出文件夹：')
        print(outputfolder + '\n')
        
    # 从文件夹中找到所有的excel
    excelsfiles = []
    count = 0
    all_file_list = os.listdir(input_path)
    for i in all_file_list:
        print(i)
        if os.path.splitext(i)[1] == '.xls':
            excel_path = input_path + '\\' + i
            output_save_path = outputfolder + '\\' + i + 'x'
            # 将xls文件转为xlsx文件,因为openpyxl不支持.xls的文件
            x = pd.read_excel(excel_path)
            x.to_excel(output_save_path, index=False)
            count += 1
        elif os.path.splitext(i)[1] == '.xlsx':
            excelsfiles.append(i)
            copy2(input_path + '\\' + i, outputfolder + '\\' + i)
            count += 1
    print('原始文件夹中xls和xlsx数量等于：' + str(count))
    excelsfiles = os.listdir(outputfolder)
    print('output文件夹中xls和xlsx数量等于：' + str(len(excelsfiles)))
    print('包括excel文件：' + str(excelsfiles) + '\n')
    return excelsfiles


def data_standarize():
    invalid_characaters = "、()（）:：，。!@#$^&*_<=,+[]{}\\;,',.?/\n)＃'%％±\xa0"
    # for i in range(firstvaluerow, maxrow+1):
    for i in range(2, maxrow + 1):
        for y in range(1, maxcolumn + 1):
            cell_value = sheet.cell(i, y).value
            # print(str(cell_value) + str(type(cell_value)) )
            if cell_value is None:
                pass
            elif type(cell_value) == str:
                string = cell_value
                for c in invalid_characaters:
                    string = string.replace(c, "")
                string = string.replace(" ", "")
                string = string.replace("　", "")
                sheet.cell(i, y).value = string
    print("已完成数据清洗")
    wb.save(excel)

# 提取数值第一行
def find_firstvalue_row():
    for row in range(1, maxrow):
        float_num = 0
        nonevalue_num = 0
        for column in range(1, maxcolumn + 1):
            s = sheet.cell(row, column).value
            # print(s, type(s), sep=',')
            if type(s) == float or type(s) == int:
                float_num += 1
            elif s is None:
                nonevalue_num += 1
        # print(float_num, nonevalue_num,sep=',')
        # 打印小数占行内所有飞空数据的比例
        if maxcolumn - nonevalue_num > 3:
            float_proportion = float_num / (maxcolumn - nonevalue_num)
            # print(float_proportion)
            if float_proportion >= 0.5:
                firstvaluerow = row
                break
    return firstvaluerow


def getfirstheadrow(firstvaluerow):
    firstheadrow = 0
    unit1 = 0
    for i in range(2, firstvaluerow):
        for y in range(1, maxcolumn):
            value = sheet.cell(i, y).value
            if type(value) == str:
                # 如果先匹配到‘单位’，则直接跳出循环，下一行为表头第一行
                match_obj = re.match("单位", value)
                if match_obj:
                    unit1 = value.split('位')[1]
                    firstheadrow = i + 1
                    break

                # 如果直接匹配到汉字，则跳出循环，此行为表头第一行
                chinese_code = ".*?([\u4E00-\u9FA5])"
                chinese_match = re.match(chinese_code, value)
                if chinese_match:
                    firstheadrow = i
                    unit1 = None
                    # print(chinese_match)
                    break
        if firstheadrow != 0:
            break
    return firstheadrow, unit1


def getheadandvalue():
    # 提取数值第一行
    firstvaluerow = find_firstvalue_row()

    # 提取表头第一行
    firstheadrow, unit = getfirstheadrow(firstvaluerow)

    # 修正数值第一行
    if firstvaluerow == firstheadrow:
        firstvaluerow += 1

    # 修正表头最后一行
    t, u = 0, 0
    for i in range(1, maxcolumn):
        cellvalue = sheet.cell(firstvaluerow - 1, i).value
        if cellvalue is None:
            t = i
            break
    m = maxcolumn - t
    if t != 0:
        for i1 in range(t, maxcolumn):
            cellvalue = sheet.cell(firstvaluerow - 1, i1).value
            if cellvalue is None:
                u += 1
            else:
                u = 0
    if u == m:
        lastheadrow = firstvaluerow - 2
    else:
        lastheadrow = firstvaluerow - 1

    print("表头从第 " + str(firstheadrow) + " 行开始")
    print("表头在第 " + str(lastheadrow) + " 行结束")
    print("值是从第 " + str(firstvaluerow) + " 行开始")
    print("单位名称：" + str(unit))

    # 合并表头
    columnnameoriginal = []
    for headcolumn in range(1, maxcolumn + 1):
        headname = []
        for headrow in range(firstheadrow, lastheadrow + 1):
            headcellvalue = sheet.cell(headrow, headcolumn).value
            headname.append(headcellvalue)
        # print(headname)
        columnnameoriginal.append(headname)
    # print(columnnameoriginal)

    i=-1
    while True:
        if columnnameoriginal[i] == [None, None]:
            del columnnameoriginal[i]
        else:
            break
    # print(columnnameoriginal)

    # 补全表头层次关系
    for r in range(1,len(columnnameoriginal)):
        headrow_num = lastheadrow - firstheadrow + 1
        w = 0
        valuer0 = columnnameoriginal[r][0]
        if valuer0 is None:
            for i in range(1, headrow_num + 1):
                valueri = columnnameoriginal[r][i]
                # print(valueri)
                if valueri is not None:
                    w = i
                    break
            for m in range(i):
                columnnameoriginal[r][m] = columnnameoriginal[r - 1][m]

    # 合并表头名称
    columnnamemerge = []
    for i in columnnameoriginal:
        list = []
        for y in range(headrow_num):
            if type(i[y]) is str:
                chinese_code = ".*?([\u4E00-\u9FA5])"
                chinese_match = re.match(chinese_code, i[y])
                number_match = re.match('\d+', i[y])
                str_match = re.findall('[a-zA-Z]', i[y])
                str_notmatch = not str_match
                if chinese_match:
                    list.append(i[y])
                elif number_match and str_notmatch:
                    list.append(i[y])
            elif type(i[y]) is int:
                list.append(str(i[y]))
        s = '_'
        mergename = s.join(list)
        if mergename in columnnamemerge:
            mergename = mergename + '1'
        columnnamemerge.append(mergename)

    # 处理只有英文的字段
    for i in range(len(columnnamemerge)):
        if len(columnnamemerge[i]) == 0:
            addname = ""
            for m in columnnameoriginal[i]:
                if type(m) == str:
                    addname = addname + m
            columnnamemerge[i] = addname

    # 提取值数据
    valuelist = []
    for i in range(firstvaluerow, maxrow + 1):
        rowlist = []
        for y in range(1, len(columnnamemerge) + 1):
            value1 = sheet.cell(i, y).value
            rowlist.append(value1)
        valuelist.append(rowlist)

     # 删除最后的备注行
    i = 0
    while i + 1:
        num = 0
        for y in range(2, maxcolumn + 1):
            # print(sheet.cell(maxrow - i, y).value)
            if sheet.cell(maxrow - i, y).value is None:
                num += 1
        # print(i, num, maxcolumn)
        if num == maxcolumn - 1:
            del valuelist[maxrow - firstvaluerow - i ]
            # print("delete" + str(i))
            i += 1
        else:
            # print("don't delete")
            break

    return columnnamemerge, valuelist


def newtable_visualise(field, valuelist,columnname_merge):
    table = PrettyTable(field)
    for i in range(len(valuelist)):
        # print(i, str(valuelist[i]))
        table.add_row(valuelist[i])
    print(table)
    print()


def createnewsheet(field, valuelist):
    if "newsheet" in wb.sheetnames:
        new_sheet = wb.get_sheet_by_name("newsheet")
        print("表格已存在")
    else:
        new_sheet = wb.create_sheet()
        # new_sheet.title = sheet_title
        new_sheet.title = "newsheet"
    alignment_style = Alignment(horizontal='center', vertical='center',wrap_text=True)
    # 写入标题
    for i in range(len(fieldname)):
        new_sheet.cell(row=1, column=i+1).value = field[i]
        new_sheet.cell(row=1, column=i+1).alignment = alignment_style
        new_sheet.cell(row=1, column=i+1).font = Font(bold=True, size=9)
    # 写入值
    for m in range(len(valuelist)):
        for n in range(len(fieldname)):
            new_sheet.cell(row=m+2, column=n + 1).value = valuelist[m][n]
    sheetnames = wb.sheetnames
    if 'Sheet1' in sheetnames:
        Sheet1 = wb.get_sheet_by_name('Sheet1')
        wb.remove(Sheet1)
    wb.save(excel)
    # wb.save(outputfolder + sheet_title + '.xlsx')

    # excel重命名
    os.rename(excel, outputfolder + sheet_title + '.xlsx')


if __name__ == '__main__':
    input_path = 'F:\\测试数据\\年鉴自动清理表头\\测试数据'
    outputfolder = 'F:\\测试数据\\年鉴自动清理表头\\output\\'

    # 获取所有年鉴excel表格并建立结果文件夹
    excels = get_all_excels()
    print("------------------开始处理excel------------------------")
    for xlsx in excels:
        excel = outputfolder + '\\' + xlsx

        # 加载数据
        wb = openpyxl.load_workbook(excel)
        sheetname = wb.sheetnames[0]  # 直接取第一张表，因为年鉴每个excel只有一张表。
        sheet = wb.get_sheet_by_name(sheetname)
        maxrow = sheet.max_row
        maxcolumn = sheet.max_column

        # 提取表标题
        sheet_title = sheet.cell(1, 1).value
        print('正在处理表： ' + sheet_title)

        # 数据标准化
        data_standarize()

        # 提取表头和数据
        fieldname, sheetvalue = getheadandvalue()

        # 可视化表格
        newtable_visualise(fieldname, sheetvalue,fieldname)

        # 新建sheet，将表格写入,删除原有sheet1
        createnewsheet(fieldname, sheetvalue)

