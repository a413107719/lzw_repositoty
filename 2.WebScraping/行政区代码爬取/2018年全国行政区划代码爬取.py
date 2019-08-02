from urllib.request import urlopen
from bs4 import BeautifulSoup
import urllib.request
import re
from time import *
import openpyxl
import os
import random
from retrying import retry


# 让函数报错后继续重新执行，达到最大执行次数的上限后才判断链接失效，不再继续执行。
@retry(stop_max_attempt_number=3)
def get_html(url):
    global curr_url
    user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36'
    response = urllib.request.Request(url)
    response.add_header('User-Agent', user_agent)
    # 对请求进行强制要求，通过添加timeout参数，让它必须在特定的时间内返回结果，否则就报错
    response = urllib.request.urlopen(response, timeout=5)
    html = BeautifulSoup(response.read(), "html.parser", from_encoding='gbk')
    return html


def get_area_list(url, kind, depth):
    try:
        html = get_html(url)
        tr_list = html.findAll('tr', {'class': 'provincetr'})
        for tr in tr_list:
            td_list = tr.findAll('td')
            for td in td_list:
                province_name = td.get_text()
                province_code = ''
                if td.a:
                    page = td.a.attrs['href']
                    index = page.index('.')
                    if index > -1:
                        substr = page[0:index]
                        province_code = substr + (12 - len(substr)) * '0'
                    pattern = re.compile(r'\w*.html')
                    province_url = re.sub(pattern, page, url)
                    print('\n' + province_name)
                    # 开始爬村镇信息
                    if kind == "all":
                        get_list(province_url, depth, province_code, province_name)
                    elif kind == province_name:
                        print('找到'+ province_name)
                        get_list(province_url, depth, province_code, province_name)
                    else:
                        print('名称不对,继续匹配。。。。。。')
                        continue

    except Exception as e:
        # 如果有出错，则回滚
        print(e)  # time out
        print('get_area_list重新请求！')
        sleep(3)
        get_area_list(url)


def get_list(provinceurl, get_level, provincecode, provincename):
    list = []
    try:
        # 获取市信息
        if get_level > 1:
            html = get_html(provinceurl)
            tr_list = html.findAll('tr', {'class': level_arr[str(2)]})
            for tr in tr_list:
                tempt_url = provinceurl
                td_list = tr.findAll('td')
                td = td_list[0]
                citytr_code = td_list[0].get_text()
                citytr_name = td_list[1].get_text()
                if td.a:
                    page = td.a.attrs['href']
                    pattern = re.compile(r'\w*.html')
                    city_url = re.sub(pattern, page, tempt_url)
                    # print(citytr_code, citytr_name, city_url)


                # 获取区县信息'
                print('\n' + "开始爬取 " + citytr_name + " 的信息")
                if get_level > 2:
                    html = get_html(city_url)
                    tempt_url = city_url
                    tr_list = html.findAll('tr', {'class':  level_arr[str(3)]})
                    for tr in tr_list:
                        td_list = tr.findAll('td')
                        td = td_list[0]
                        countytr_code = td_list[0].get_text()
                        countytr_name = td_list[1].get_text()
                        if td.a:
                            page = td.a.attrs['href']
                            pattern = re.compile(r'\w*.html')
                            countytr_url = re.sub(pattern, page, tempt_url)
                            # print(citytr_code,citytr_name,countytr_code,countytr_name,countytr_url,sep=' ')

                            # 获取乡镇信息
                            if get_level > 3:
                                html = get_html(countytr_url)
                                tempt_url2 = countytr_url
                                tr_list = html.findAll('tr', {'class': level_arr[str(4)]})
                                for tr in tr_list:
                                    td_list = tr.findAll('td')
                                    td = td_list[0]
                                    town_code = td_list[0].get_text()
                                    town_name = td_list[1].get_text()
                                    if td.a:
                                        page = td.a.attrs['href']
                                        pattern = re.compile(r'\w*.html')
                                        town_url = re.sub(pattern, page, tempt_url2)
                                        # print(citytr_code, citytr_name, countytr_code, countytr_name, town_code, town_name,town_url, sep=' ')

                                        # 获取村、社区信息
                                        if get_level > 4:
                                            html = get_html(town_url)
                                            tempt_url3 = town_url
                                            tr_list = html.findAll('tr', {'class': level_arr[str(5)]})
                                            # print(tr_list)
                                            for tr in tr_list:
                                                td_list = tr.findAll('td')
                                                td = td_list[0]
                                                village_code = td_list[0].get_text()
                                                village_number = td_list[1].get_text()
                                                village_name = td_list[2].get_text()
                                                if td.a:
                                                    page = td.a.attrs['href']
                                                    pattern = re.compile(r'\w*.html')
                                                    village_url = re.sub(pattern, page, tempt_url3)
                                                    # print(provincecode, provincename, citytr_code, citytr_name, countytr_code, countytr_name, town_code, town_name, village_code, village_number, village_name, village_url, sep=' ')
                                                else:
                                                    villagelist = [provincecode, provincename, citytr_code, citytr_name, countytr_code, countytr_name, town_code, town_name, village_code, village_number, village_name]
                                                    print(villagelist)
                                                    list.append(villagelist)
                                               
                                    else:
                                        townlist= [provincecode, provincename, citytr_code, citytr_name, countytr_code, countytr_name, town_code, town_name]
                                        print(townlist)
                                        list.append(townlist)
                                    sleep(random.randint(1, 4))

                        else:
                            countytrlist = [provincecode, provincename, citytr_code, citytr_name, countytr_code, countytr_name]
                            print(countytrlist)
                            list.append(countytrlist)
                        sleep(random.randint(2, 5))

        print(provincename + "的信息爬取完毕")
        # 写入excel
        write2excel(list, provincename)

    except Exception as e:
        # 如果有出错，则回滚
        print(e)  # time out
        print('出错，重新请求！')
        sleep(3)
        get_list(provinceurl, get_level, provincecode, provincename)


def write2excel(list, provincename):
    excelname = '行政区划代码爬取'
    outputfolder = 'D:\\'
    excelpath = outputfolder + excelname + '.xlsx'
    if not os.path.isdir(excelpath):
        print("创建 行政区划代码爬取.xlsx")
        wb = openpyxl.Workbook()
        wb.save(excelpath)

    wb = openpyxl.load_workbook(excelpath)
    wb.create_sheet(provincename)
    sheet = wb[provincename]
    headnamelist = ["省行政区划代码", "省名", "市行政区划代码", "市名", "区县行政区划代码", "区县名", "镇行政区划代码", "镇名", "村社区行政区划代码", "城乡分类代码", "村社区名称"]
    for i in range(len(headnamelist)):
        sheet.cell(row=1, column=i+1).value = headnamelist[i]
    for i in range(len(list)):
        for x in range(len(list[i])):
            cellvalue = list[i][x]
            sheet.cell(row=i+2, column=x+1).value = cellvalue
    wb.save(excelpath)
    print("Finish，已写入excel")


if __name__ == '__main__':
    '''
    函数get_area_list(url, kind, depth)中    
    depth参数：1省，2市，3区县，4，乡镇，5村或社区
    kind参数:all爬取全国所有省份数据，"重庆市"只爬取重庆市的数据
    省名包括：[北京市,天津市,河北省,山西省,内蒙古自治区,辽宁省,吉林省,黑龙江省,上海市,江苏省,
    浙江省,安徽省,福建省,江西省,山东省,河南省,湖北省,湖南省,广东省,广西壮族自治区,海南省,
    重庆市,四川省,贵州省,云南省,西藏自治区,陕西省,甘肃省,青海省,宁夏回族自治区,新疆维吾尔自治区,]
    
    '''

    level_arr = {'1': 'provincetr', '2': 'citytr', '3': 'countytr', '4': 'towntr', '5': 'villagetr'}
    url = 'http://www.stats.gov.cn/tjsj/tjbz/tjyqhdmhcxhfdm/2018/index.html'
    excel_path = 'D:\\'
    curr_url = ''
    get_area_list(url, "广西壮族自治区", 5)
