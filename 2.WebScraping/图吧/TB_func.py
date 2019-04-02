import requests
import re
import time
import openpyxl
import os
from bs4 import BeautifulSoup


# 找到每个点的信息
def point_data(point_url):
    web_data = requests.get(point_url)
    soup = BeautifulSoup(web_data.text,'lxml')
    names = soup.find_all('meta')
    text = names[2].get('content')       #province=四川;city=自贡;coord=104.76247,29.37363
    textregex = re.compile(r'province=(\S\S);city=(\S\S);coord=(\d+\W\d+),(\d+\W\d+)')
    mo = textregex.search(text)
    province,city,lattitude,longitude = mo.groups()
    # print(province,city,lattitude,longitude,sep='\n')
    return province,city,lattitude,longitude

# 打包信息
def pack_imformation(each_urlANDkinds_name):
    # url = 'http://poi.mapbar.com/zigong/FC0/'
    each_url = str(each_urlANDkinds_name).split(',')[0]
    kinds_name = str(each_urlANDkinds_name).split(',')[1]
    web_data = requests.get(each_url)
    soup = BeautifulSoup(web_data.text,'lxml')
    names = soup.select('body > div.sort.cl > div.sortC > dl > dd > a')
    links = soup.select('body > div.sort.cl > div.sortC > dl > dd > a')
    # print(names)
    # 设置表格路径
    os.chdir('C:\\Users\\lzw\\Desktop')
    wb = openpyxl.load_workbook('example.xlsx')
    if "地名爬取" in str(wb.get_sheet_by_name("地名爬取")):
        pass
    else:
        wb.create_sheet('地名爬取')  # 新建一个excel页面
    sheet = wb['地名爬取']                   # 选择一个excel页面
    ID = 1
    for name, link in zip(names, links):
        point_url = link.get('href')
        try:
            province, city, lattitude, longitude = point_data(point_url) #找到每个点信息
            # 表格填入数据
            max_row = sheet.max_row
            newrow = max_row + 1
            sheet.cell(row=newrow, column=1).value = ID
            sheet.cell(row=newrow, column=2).value = name.string
            sheet.cell(row=newrow, column=3).value = province
            sheet.cell(row=newrow, column=4).value = city
            sheet.cell(row=newrow, column=5).value = lattitude
            sheet.cell(row=newrow, column=6).value = longitude
            sheet.cell(row=newrow, column=7).value = kinds_name
            wb.save('example.xlsx')
            print('已经下载数据：' + name.string + '  '+ kinds_name)
            ID += 1
        except:
            pass


each_urlANDkinds_name = 'http://poi.mapbar.com/zigong/BD0/,自贡市其他公司企业'
pack_imformation(each_urlANDkinds_name)