import openpyxl
import os

def getmaxrow(sheet):
    row = 3
    while True:
        if sheet.cell(row=row, column=3).value is None:
            break
        else:
            row += 1
    max_row = row - 1
    max_column = sheet.max_column
    return max_row, max_column


def getsheetsname(sheet):
    x = 1
    while sheet.cell(row=1, column=x).value != '空间位置':
        x += 1

    # 得到数据名称列表
    sheetnamelist = []
    namelist=[]
    for y in range(3, maxrow+1):
        list = []
        for i in range(1, x):
            cell = sheet.cell(row=y, column=i).value
            if cell is None:
                cell = namelist[-1][i-1]
            list.append(cell)
        namelist.append(list)
    for y in namelist:
        sheetnamelist.append('_'.join(i for i in y))

    # 得到字段列表
    sheetattributelist = []
    for y in range(3, maxrow+1):
        list = []
        for i in range(x+1, maxcolumn):
            cell = sheet.cell(row=y, column=i).value
            if not cell is None:
                list.append(cell)
        sheetattributelist.append(list)
    return sheetnamelist,  sheetattributelist


def getsheetattributes():
    attributeslist = []


    print(attributeslist)
    return attributeslist


def write2excel(namelist, attributelist):
    # 新建excel，以sheetname命名
    excelpath = folder + sheetname + '.xlsx'
    wb = openpyxl.Workbook()
    for i in range(len(namelist)):
        wbs = wb.create_sheet()
        wbs.title = namelist[i]
        wbattributes = attributelist[i]
        for y in range(len(wbattributes)):
            wbs.cell(row=1, column=y+1).value = wbattributes[y]

    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    wb.save(excelpath)
    print("Finish，已写入excel")


folder = "D:\\基础数据库\\"
xlsx = folder + "基础数据库设计.xlsx"
wb = openpyxl.load_workbook(xlsx)
sheets = ["县域维度数据", "县域事实数据", "中心城区维度数据", "中心城区事实数据"]
# for sheetname in sheets:
sheetname = '县域维度数据'  # 注释掉
sheet = wb[sheetname]
maxrow, maxcolumn = getmaxrow(sheet)
namelist, attributelist = getsheetsname(sheet)  # 构造表名和字段列表
write2excel(namelist, attributelist)  # 创建excel并写入字段名称
