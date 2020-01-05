# coding=utf-8
import re
from prettytable import PrettyTable
from openpyxl.styles import Color, Fill, Font, Alignment, PatternFill, Border, Side
import openpyxl
import pandas as pd
import os
import shutil
from shutil import copy2
from pandas import DataFrame
import pandas as pd



def data_standarize(list):
    newlist = []
    invalid_characaters = "、()（）:：，。!@#$^&*_<=,+[]{}\\;,',.?/\n)＃'%％±\xa0"
    for strings in list:
        for c in invalid_characaters:
            strings = strings.replace(c, "")
        strings = strings.replace(" ", "")
        strings = strings.replace("　", "")
        newlist.append(strings)
    print("已完成特殊字符清洗")
    return newlist


# 提取数值第一行
def find_firstvalue_row():
    firstvaluerow = 0
    for row in range(2, maxrow):
        float_num = 0
        nonevalue_num = 0
        columnvalue_list=[]
        for column in range(1, maxcolumn + 1):
            s = sheet.cell(row, column).value
            columnvalue_list.append(s)
        while True:
            # print('colunmnvalu_list: ' + str(columnvalue_list))
            if columnvalue_list == []:
                break
            elif columnvalue_list[-1] is None:
                del columnvalue_list[-1]
            else:
                break
        # print(columnvalue_list)

        # 避免excel中数字显示为文本，导致无法统计
        newlist = []
        for i in columnvalue_list:
            if type(i) == int or type(i) == float or i is None:
                newlist.append(i)
            elif i.isdigit():
                newlist.append(int(i))
            else:
                newlist.append(i)
        # print(newlist)

        for s in newlist:
            if type(s) == float or type(s) == int:
                float_num += 1
            elif s is None:
                nonevalue_num += 1
        # 打印小数占行内所有非空数据的比例
        # print(maxcolumn,nonevalue_num)
        if newlist == []:
            pass
        elif maxcolumn - nonevalue_num >= 2:
            float_proportion = float_num / (len(newlist) - nonevalue_num)
            # print(row, maxcolumn, float_num, nonevalue_num,float_proportion)
            if float_proportion >= 0.2:
                firstvaluerow = row
                break
    # print('firstvaluerow:' + str(firstvaluerow))
    return firstvaluerow


def getfirstheadrow(firstvalue_row):
    firstheadrow = 0
    unit1 = 0
    print("firstvaluerow:" + str(firstvalue_row))
    for i in range(2, firstvalue_row):
        for y in range(1, maxcolumn+1):
            value = sheet.cell(i, y).value
            # 排除字段名称为"单位"
            list = [value, sheet.cell(i, y + 1).value, sheet.cell(i, y + 2).value, sheet.cell(i, y + 3).value]
            # print('list:' + str(list))
            while True:
                if list == []:
                    break
                elif list[-1] is None:
                    del list[-1]
                else:
                    break
            if type(value) == str:
                # 匹配包含‘单位’的字段
                match_obj = re.match("单位", value)
                if match_obj and len(list) >= 2:
                    firstheadrow = i
                    break
                elif match_obj:
                    unit1 = value.split('位')[1]
                    firstheadrow = i + 1
                    break

                # 如果直接匹配到汉字，则跳出循环，此行为表头第一行
                chinese_code = ".*?([\u4E00-\u9FA5])"
                chinese_match = re.match(chinese_code, value)
                if chinese_match:
                    firstheadrow = i
                    unit1 = None
                    break
            elif value is None:
                continue
        if firstheadrow != 0:
            break
    print(firstheadrow)
    return firstheadrow, unit1


def getheadandvalue():
    # 提取数值第一行
    firstvaluerow = find_firstvalue_row()
    # print('firstvaluerow' + str(firstvaluerow))

    # 提取表头第一行
    firstheadrow, unit = getfirstheadrow(firstvaluerow)
    # print('firstheadrow' + str(firstheadrow))


    # 修正数值第一行
    if firstvaluerow == firstheadrow:
        firstvaluerow += 1

    # 修正表头最后一行
    t, u = 0, 0
    for i in range(1, maxcolumn):
        cellvalue = sheet.cell(firstvaluerow - 1, i).value
        if cellvalue is None:
            t = i
            break
    m = maxcolumn - t
    if t != 0:
        for i1 in range(t, maxcolumn):
            cellvalue = sheet.cell(firstvaluerow - 1, i1).value
            if cellvalue is None:
                u += 1
            else:
                u = 0
    if u == m:
        lastheadrow = firstvaluerow - 2
    else:
        lastheadrow = firstvaluerow - 1

    print("表头从第 " + str(firstheadrow) + " 行开始")
    print("表头在第 " + str(lastheadrow) + " 行结束")
    print("值是从第 " + str(firstvaluerow) + " 行开始")
    print("单位名称：" + str(unit))

    # 合并表头
    columnnameoriginal = []
    for headcolumn in range(1, maxcolumn + 1):
        headname = []
        for headrow in range(firstheadrow, lastheadrow + 1):
            headcellvalue = sheet.cell(headrow, headcolumn).value
            headname.append(headcellvalue)
        # print(headname)
        columnnameoriginal.append(headname)
    # print(columnnameoriginal)

    i=-1
    while True:
        if columnnameoriginal[i] == [None, None]:
            del columnnameoriginal[i]
        else:
            break
    # print('columnnameoriginal: ' + str(len(columnnameoriginal)))
    print(columnnameoriginal)

    # 补全表头层次关系
    if len(columnnameoriginal[1]) == 1:
        while True:
            if columnnameoriginal[-1][0] is None:
                del columnnameoriginal[-1]
            else:
                break
    headrow_num = lastheadrow - firstheadrow + 1
    for r in range(1,len(columnnameoriginal)-1):
        # headrow_num = lastheadrow - firstheadrow + 1
        w = 0
        # print(len(columnnameoriginal[1]),r,columnnameoriginal[r][0])
        if columnnameoriginal[r][0] is None:
            for i in range(1, headrow_num + 1):
                # print(r, i, headrow_num)
                valueri = columnnameoriginal[r][i]
                # print(valueri)
                if valueri is not None:
                    w = i
                    break
            for m in range(i):
                columnnameoriginal[r][m] = columnnameoriginal[r - 1][m]

    # 合并表头名称
    columnnamemerge = []
    for i in columnnameoriginal:
        list = []
        for y in range(headrow_num):
            if type(i[y]) is str:
                chinese_code = ".*?([\u4E00-\u9FA5])"
                chinese_match = re.match(chinese_code, i[y])
                number_match = re.match('\d+', i[y])
                str_match = re.findall('[a-zA-Z]', i[y])
                str_notmatch = not str_match
                if chinese_match:
                    list.append(i[y])
                elif number_match and str_notmatch:
                    list.append(i[y])
            elif type(i[y]) is int:
                list.append(str(i[y]))
        s = '_'
        mergename = s.join(list)
        if mergename in columnnamemerge:
            mergename = mergename + '1'
        columnnamemerge.append(mergename)

    # 处理只有英文的字段
    for i in range(len(columnnamemerge)):
        if len(columnnamemerge[i]) == 0:
            addname = ""
            for m in columnnameoriginal[i]:
                if type(m) == str:
                    addname = addname + m
            columnnamemerge[i] = addname

    # 提取值数据
    valuelist = []
    for i in range(firstvaluerow, maxrow + 1):
        rowlist = []
        for y in range(1, len(columnnamemerge) + 1):
            value1 = sheet.cell(i, y).value
            rowlist.append(value1)
        valuelist.append(rowlist)

     # 删除最后的备注行
    i = 0
    while i + 1:
        num = 0
        for y in range(1, maxcolumn + 1):
            if sheet.cell(maxrow - i, y).value is None:
                num += 1
        if num == maxcolumn - 1:
            del valuelist[maxrow - firstvaluerow - i ]
            i += 1
        else:
            break

    # 清理字段名称
    columnnamemerge = data_standarize(columnnamemerge)
    return columnnamemerge, valuelist


def newtable_visualise(field, valuelist,columnname_merge):
    print(field)
    table = PrettyTable(field)
    for i in range(len(valuelist)):
        # print(i, str(valuelist[i]))
        table.add_row(valuelist[i])
    print(table)
    print()



def createnewsheet(field, valuelist, new_sheetname,excelfile):
    wb1 = openpyxl.load_workbook()
    new_sheet = wb1.create_sheet(new_sheetname)
    # 写入标题
    alignment_style = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for i in range(len(fieldname)):
        new_sheet.cell(row=1, column=i+1).value = field[i]
        new_sheet.cell(row=1, column=i+1).alignment = alignment_style
        new_sheet.cell(row=1, column=i+1).font = Font(bold=True, size=9)
    # 写入值
    for m in range(len(valuelist)):
        for n in range(len(fieldname)):
            new_sheet.cell(row=m+2, column=n + 1).value = valuelist[m][n]
    wb1.save(excelfile)
    print("success")




excel = 'F:\\测试数据\\年鉴自动清理表头\\测试数据\\2019酉阳统计年鉴数据汇总修改.xlsx'
out_path = 'F:\\测试数据\\年鉴自动清理表头\\output'

# 加载数据
wb = openpyxl.load_workbook(excel)
sheetnames = wb.sheetnames  # 直接取第一张表，因为年鉴每个excel只有一张表。
print(sheetnames, len(sheetnames))
errorsheet = []
sheettitles = []
for originalsheetname in sheetnames:
    try:
        print('开始处理：', originalsheetname)
        sheet = wb[originalsheetname]
        maxrow = sheet.max_row
        maxcolumn = sheet.max_column
        print(maxrow,maxcolumn)

        # 提取表标题
        newsheetname = data_standarize([sheet.cell(1, 1).value])[0]
        sheettitles.append(newsheetname)

        # 提取表头和数据
        fieldname, sheetvalue = getheadandvalue()

        # 可视化表格
        newtable_visualise(fieldname, sheetvalue, fieldname)

        # file_path = out_path + '\\' + newsheetname + '.xlsx'
        # createnewsheet(fieldname, sheetvalue, newsheetname,file_path)
        # print('sgfdstertew')

        df = DataFrame(sheetvalue,columns=fieldname)
        print(df)
        csv_path = out_path + '\\' +newsheetname + '.csv'
        df.to_csv(csv_path,index=False, encoding="utf_8_sig")

        # file_path = out_path + '\\' +newsheetname + '.xlsx'
        # print(file_path)
        # df.to_excel(file_path,encoding='utf-8',index=False,header=True)
        # print("success")

    except Exception as e:
        errorsheet.append(originalsheetname)
        continue
