import arcpy
import os
import openpyxl
import time
import socket
import uuid


# 新建一个output数据库
def create_gdb(input_folder):
    if arcpy.Exists("ProjectGDB.gdb"):
        print('ProjectGDB.gdb数据库已存在,将直接把数据写入数据库')
    else:
        arcpy.CreateFileGDB_management(input_folder, "ProjectGDB.gdb")
        print("成功创建ProjectGDB.gdb数据库")


# 提取点、线、面和注记数据
def dwg2gdb(input_folder, output_folder, muban_folder):
    wb = openpyxl.load_workbook(muban_folder + '地形图提取模板.xlsx')
    os.chdir(input_folder)
    cads = arcpy.ListFiles("*.dwg")
    print('文件夹中存在CAD数据：' + str(cads) + '\n')

    # 提取数据
    feature_kinds = ['Point', 'Polyline', 'Polygon', 'Annotation']
    for feature_kind in feature_kinds:
        sheet = wb[feature_kind]
        max_row = sheet.max_row
        print("共有" + str(max_row - 1) + "条" + feature_kind + "数据将会被提取出来")
        for cad in cads:
            print('【正在从"' + cad + '"中提取' + feature_kind + '数据】')
            dwg_feature = input_folder + cad + "\\" + feature_kind
            arcpy.env.workspace = output_folder
            arcpy.env.overwriteOutput = True
            for name in range(2, max_row+1):
                feature_name = sheet.cell(name, 1).value + '_' + cad.split('.')[0]
                selet_condition = sheet.cell(name, 2).value
                if feature_name is None:
                    pass
                else:
                    print('[' + feature_name + ']')
                    print(dwg_feature)
                    print(output_folder)
                    print(feature_name)
                    print(selet_condition)
                    arcpy.FeatureClassToFeatureClass_conversion(dwg_feature,
                                                                output_folder, feature_name, selet_condition)
                    print("成功在数据库中添加要素：" + feature_name + '\n')
            print('-----------------------' + '\n' + '【' + feature_kind + "数据已提取完成" + '】' + "\n")

    # 删除空数据
    print('【' + "正在删除空数据" + '】')
    arcpy.env.workspace = output_folder
    for shpfile in arcpy.ListFeatureClasses():
        rowcount = arcpy.GetCount_management(shpfile)
        print("数据名称： " + shpfile + "     数据量：" + str(rowcount))
        if int(str(rowcount)) == 0:
            fullpath = output_folder + '\\' + shpfile
            arcpy.Delete_management(fullpath)  # 在GDB中删除数据
            print(shpfile + "： 由于数据为空，已在GDB中删除")
    print("【" + "CAD数据入库已完成" '】' + "\n")


# 新建map + 添加数据 + 可视化
def data_visualization(input_folder, output_folder, muban_folder):
    pro_aprx = arcpy.mp.ArcGISProject(muban_folder + "模板.aprx")

    # 在map中添加数据
    print('【' + '开始在map中添加数据' + '】')
    data_map = pro_aprx.listMaps("Map")[0]
    arcpy.env.workspace = output_folder
    for feature_name in arcpy.ListFeatureClasses():
        feature_location = output_folder + "\\" + feature_name
        data_map.addDataFromPath(feature_location)
        print("成功添加数据: " + feature_name)
    print('【' + "数据添加完成" + '】' + '\n')
    data_layers = data_map.listLayers()

    # 设置数据样式
    print('【' + '开始设置符号样式' + '】')
    wb = openpyxl.load_workbook(muban_folder + '地形图提取模板.xlsx')
    os.chdir(input_folder)
    feature_kinds = ['Point', 'Polyline', 'Polygon', 'Annotation']
    for feature_kind in feature_kinds:
        sheet = wb[feature_kind]
        max_row = sheet.max_row
        for row_num in range(2, max_row+1):
            row_name = sheet.cell(row_num, 1).value
            for layer in data_layers:
                layer_name = str(layer)
                if layer_name.split('_')[0] == row_name:           # 根据excel表在GDB中选择相应的数据做符号化和标注设置
                    print('[' + '开始设置 "' + layer_name + '" 的符号和标注样式' + ']')
                    sheet_contain_field = sheet.cell(row_num, 3).value       # 保留字段
                    sheet_lable_size = sheet.cell(row_num, 4).value          # 设置size
                    sheet_color_way = sheet.cell(row_num, 5).value           # 渲染方式
                    sheet_color_field = sheet.cell(row_num, 6).value         # 渲染字段
                    sheet_color_ramp = sheet.cell(row_num, 7).value          # 色带和符号库样式
                    lable_ornot = sheet.cell(row_num, 8).value               # 标注与否
                    lable_field = sheet.cell(row_num, 9).value               # 标注字段

                    # 清理数据字段
                    if sheet_contain_field == ' ' or sheet_contain_field is None:
                        print('数据字段未清理')
                    else:
                        remain_field_list = sheet_contain_field.split(',') + ['OBJECTID', 'Shape',
                                                                              'Shape_Length', 'Shape_area']
                        print("保留字段" + str(remain_field_list))
                        # 生成字段名列表
                        original_field_list = [f.name for f in arcpy.ListFields(layer)]
                        print('数据原始字段包括：' + str(original_field_list))
                        for i in remain_field_list:
                            if i in original_field_list:
                                original_field_list.remove(i)
                            else:
                                pass
                        delete_field_list = original_field_list
                        if len(delete_field_list) != 0:
                            print('删除字段包括：' + str(delete_field_list))
                            arcpy.DeleteField_management(layer, delete_field_list)
                        else:
                            pass
                        print('数据字段已清理')

                    # 设置符号
                    sym = layer.symbology
                    if sheet_color_way == 'GraduatedColorsRenderer':
                        sym.updateRenderer(sheet_color_way)
                        sym.renderer.classificationField = sheet_color_field
                        sym.renderer.breakCount = 20
                        sym.renderer.colorRamp = pro_aprx.listColorRamps(sheet_color_ramp)[0]
                    elif sheet_color_way == 'UniqueValueRenderer':
                        sym.updateRenderer(sheet_color_way)
                        sym.renderer.fields = [sheet_color_field]
                        sym.renderer.colorRamp = pro_aprx.listColorRamps(sheet_color_ramp)[0]
                    else:
                        sym.renderer.symbol.applySymbolFromGallery(sheet_color_ramp)
                        # 获取并设置点符号大小、线符号宽度或面符号的轮廓宽度。
                        sym.renderer.symbol.size = sheet_lable_size
                    layer.symbology = sym
                    print('数据符号已设置')

                    # 设置标注
                    if layer.supports("SHOWLABELS"):
                        if lable_ornot == '是':
                            lblclass = layer.listLabelClasses()[0]
                            lblclass.expression = '$feature.' + str(lable_field)
                            lblclass.visible = True
                            layer.showLabels = True
                            print("图层标注已打开" + '\n')
                        else:
                            print("图层标注未设置" + '\n')
                    else:
                        pass
                else:
                    pass
    print('【' + '符号样式设置已完成' + '】' + '\n')

    # 另存为新的项目文件
    time_text = time.strftime('%Y%m%d_%H%M', time.localtime(time.time()))
    file_name = "地形图信息提取" + time_text + ".aprx"
    pro_aprx.saveACopy(input_folder + file_name)
    print('地形图信息提取完成，并已另存为aprx文件：' + file_name)
    del pro_aprx


# 身份验证
def compare_pcinfo():
    pc_list = [['zlk', '74:D0:2B:9E:4A:0A', '192.168.3.253'], ['zmlpc', '9C:5C:8E:C1:E9:4B', '192.168.3.243'],
               ['ltpc', '30:5A:3A:78:C9:F6', '192.168.3.251'], ['lxpc', '30:5A:3A:82:09:A6', '192.168.3.14'],
               ['lxpc', '30:B4:9E:54:EF:C8', '192.168.3.157'],
               ['30-B4-9E-54-EE-83', '30-B4-9E-54-EE-83', '192.168.3.177'],
               ['00-11-22-CA-DE-1C', '00:11:22:CA:DE:1C', '192.168.3.242'],
               ['server', '14:18:77:3A:C6:90', '192.168.3.16'],
               ['sylvia', 'E0:3F:49:A6:CA:80', '192.168.3.215'],
               ['2013-20170325NM', '30:B4:9E:54:EF:89', '192.168.3.232'],
               ['2013-20170325NM', '2013-20170325NM', '192.168.3.226'],
               ['christine', '10:BF:48:B9:E2:23', '192.168.3.252'],
               ['jnPC', '14:DA:E9:C7:59:9E', '192.168.3.250'], ['研究中心电脑', '3c:f8:62:eb:8d:08', '192.168.199.160']]
    print('【此电脑的基础信息包括：】')
    mac = uuid.UUID(int=uuid.getnode()).hex[-12:]
    mac_address = ":".join([mac[e:e+2] for e in range(0, 11, 2)])
    hostname = socket.gethostname()
    result = socket.getaddrinfo(hostname, None, 0, socket.SOCK_STREAM)
    ip_address = [x[4][0] for x in result][1]
    print('主机名：' + hostname)
    print('IP地址：' + ip_address)
    print('mac地址：' + mac_address)
    for row in pc_list:
        if row[1] == mac_address and row[2] == ip_address:
            print("身份验证通过！" + '\n')
            return 'yes'
    print('ZK工作室版权所有，如需使用请联系：lzw19910@qq.com')


if __name__ == '__main__':
    # 身份验证
    if compare_pcinfo() == 'yes':
        pass
    else:
        exit()

    # 独立运行版本
    cad_folder = 'F:\\测试数据\\地形图信息提取\\'
    template_folder = 'F:\\测试数据\\模板\\'

    # GIS Pro插件版本
    # cad_folder = arcpy.GetParameterAsText(0) + '\\'
    # template_folder = arcpy.GetParameterAsText(1) + '\\'

    output = cad_folder + "ProjectGDB.gdb"
    arcpy.env.workspace = cad_folder
    # 新建一个output数据库
    create_gdb(cad_folder)
    # 找到所有dwg文件，并提取信息到数据库
    dwg2gdb(cad_folder, output, template_folder)
    # 新建aprx工程文件，新建map，数据可视化
    data_visualization(cad_folder, output, template_folder)