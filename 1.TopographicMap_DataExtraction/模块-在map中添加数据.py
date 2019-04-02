import arcpy
import os
import openpyxl
import time


# 新建map + 添加数据 + 可视化
def data_visualization():
    pro_aprx = arcpy.mp.ArcGISProject(r"F:\测试数据\地形图信息提取\testproject\testproject1.aprx")

    # 在map中添加数据
    data_map = pro_aprx.listMaps("Map")[0]
    arcpy.env.workspace = output_folder
    for feature_name in arcpy.ListFeatureClasses():
        feature_location = output_folder + "\\" + feature_name
        data_map.addDataFromPath(feature_location)
        print("成功添加数据: " + feature_name)
    print("数据添加完成" + '\n')
    data_layers = data_map.listLayers()

    # 设置数据样式
    os.chdir(input_folder)
    wb = openpyxl.load_workbook('地形图提取工具模板.xlsx')
    feature_kinds = ['Point', 'Polyline', 'Polygon', 'Annotation']
    for feature_kind in feature_kinds:
        sheet = wb[feature_kind]
        max_row = sheet.max_row
        for row_num in range(2, max_row+1):
            row_name = sheet.cell(row_num, 1).value
            for layer in data_layers:
                layer_name = str(layer)
                if layer_name == row_name:           # 根据excel表在GDB中选择相应的数据做符号化和标注设置
                    print("开始设置 " + layer_name + " 的符号和标注样式。")
                    sheet_contain_field = sheet.cell(row_num, 3).value       # 保留字段
                    sheet_lable_size = sheet.cell(row_num, 4).value          # 设置size
                    sheet_color_way = sheet.cell(row_num, 5).value           # 渲染方式
                    sheet_color_field = sheet.cell(row_num, 6).value         # 渲染字段
                    sheet_color_ramp = sheet.cell(row_num, 7).value          # 色带和符号库样式
                    lable_ornot = sheet.cell(row_num, 8).value               # 标注与否
                    lable_field = sheet.cell(row_num, 9).value               # 标注字段

                    # 清理数据字段
                    if sheet_contain_field == ' ' or sheet_contain_field is None:
                        print('没有可清理的字段')
                    else:
                        remain_field_list = sheet_contain_field.split(',') + ['OBJECTID', 'Shape',
                                                                              'Shape_Length', 'Shape_area']
                        print("保留字段" + str(remain_field_list))
                        # 生成字段名列表
                        original_field_list = [f.name for f in arcpy.ListFields(layer)]
                        print('数据原始字段包括：' + str(original_field_list))
                        for i in remain_field_list:
                            if i in original_field_list:
                                original_field_list.remove(i)
                            else:
                                pass
                        delete_field_list = original_field_list
                        if len(delete_field_list) != 0:
                            print('删除字段包括：' + str(delete_field_list))
                            arcpy.DeleteField_management(layer, delete_field_list)
                        else:
                            pass
                        print('数据字段已完成清理' + '\n')

                    # 设置符号
                    sym = layer.symbology
                    if sheet_color_way == 'GraduatedColorsRenderer':
                        sym.updateRenderer(sheet_color_way)
                        sym.renderer.classificationField = sheet_color_field
                        sym.renderer.breakCount = 20
                        sym.renderer.colorRamp = pro_aprx.listColorRamps(sheet_color_ramp)[0]
                    elif sheet_color_way == 'UniqueValueRenderer':
                        sym.updateRenderer(sheet_color_way)
                        sym.renderer.fields = [sheet_color_field]
                        sym.renderer.colorRamp = pro_aprx.listColorRamps(sheet_color_ramp)[0]
                    else:
                        sym.renderer.symbol.applySymbolFromGallery(sheet_color_ramp)
                        # 获取并设置点符号大小、线符号宽度或面符号的轮廓宽度。
                        sym.renderer.symbol.size = sheet_lable_size
                    layer.symbology = sym

                    # 设置标注
                    if layer.supports("SHOWLABELS"):
                        if lable_ornot == '是':
                            lblclass = layer.listLabelClasses()[0]
                            lblclass.expression = '$feature.' + str(lable_field)
                            lblclass.visible = True
                            layer.showLabels = True
                            print("已为图层打开标注")
                        else:
                            pass
                    else:
                        print("图层未打开标注" + '\n')
                else:
                    pass

    # 另存为新的项目文件
    time_text = time.strftime('%Y%m%d_%H%M', time.localtime(time.time()))
    file_name = "地形图信息提取" + time_text + ".aprx"
    pro_aprx.saveACopy(input_folder + file_name)
    print('已另存为文件：' + file_name)
    del pro_aprx


input_folder = 'F:\\测试数据\\地形图信息提取\\'  # 以后替换为 Input_folder = arcpy.GetParameterAsText(0)
output_folder = input_folder + "ProjectGDB.gdb"
data_visualization()
