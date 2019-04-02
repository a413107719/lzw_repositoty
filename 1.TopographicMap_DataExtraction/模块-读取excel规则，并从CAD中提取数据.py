import openpyxl
import os
import arcpy


# 提取点、线、面和注记数据
def dwg2gdb():
    os.chdir(input_folder)
    cads = arcpy.ListFiles("*.dwg")
    print('文件夹中存在的CAD数据包括：' + '\n' + str(cads) + '\n')
    wb = openpyxl.load_workbook('地形图提取工具模板.xlsx')

    # 提取数据
    feature_kinds = ['Point', 'Polyline', 'Polygon', 'Annotation']
    for feature_kind in feature_kinds:
        sheet = wb[feature_kind]
        max_row = sheet.max_row
        print("共有" + str(max_row - 1) + "条" + feature_kind + "数据将会被提取出来")
        for cad in cads:
            print('正在从"' + cad + '"中提取' + feature_kind + '数据')
            dwg_feature = input_folder + cad + "\\" + feature_kind
            arcpy.env.workspace = output_folder
            for name in range(2, max_row+1):
                # 有相同文件名自动添加后缀
                feature_name = sheet.cell(name, 1).value
                i = 1
                while True:
                    if arcpy.Exists(feature_name):
                        print("数据库中存在数据：%s" % feature_name)
                        feature_name = sheet.cell(name, 1).value + '_' + str(i)
                        i += 1
                        continue
                    else:
                        break
                arcpy.FeatureClassToFeatureClass_conversion(dwg_feature, output_folder, feature_name, sheet.cell(name, 2).value)
                print("成功在数据库中添加要素：" + feature_name)
            print(feature_kind + "数据已提取完成" + "\n")

    # 删除空数据
    arcpy.env.workspace = output_folder
    for shp in arcpy.ListFeatureClasses():
        rowcount = arcpy.GetCount_management(shp)
        print("数据名称： " + shp + "     数据量：" + str(rowcount))
        if int(str(rowcount)) == 0:
            fullpath = output_folder + '\\' + shp
            arcpy.Delete_management(fullpath)  # 在GDB中删除数据
            print(shp + "： 由于数据为空，已在GDB中删除")
    print("\n" + "CAD数据入库已完成" + "\n")


input_folder = 'F:\\测试数据\\地形图信息提取\\'     # 以后替换为 input_folder = arcpy.GetParameterAsText(0)
output_folder = input_folder + "ProjectGDB.gdb"
arcpy.env.workspace = input_folder
dwg2gdb()
#test