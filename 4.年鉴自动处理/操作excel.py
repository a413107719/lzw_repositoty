import openpyxl,os
from openpyxl import Workbook

#
# os.chdir('C:\\Users\\lzw19\\Desktop')
# wb = openpyxl.load_workbook('example.xlsx')
# sheet = wb['Sheet1']         # 选择一个excel页面
# # 填入数据
# max_row = sheet.max_row
# newrow = max_row + 1
# sheet.cell(row=newrow, column=1).value = 'ID'
# sheet.cell(row=newrow, column=2).value = 'name.string'
# sheet.cell(row=newrow, column=3).value = 'province'
# sheet.cell(row=newrow, column=4).value = 'city'
# sheet.cell(row=newrow, column=5).value = 'lattitude'
# sheet.cell(row=newrow, column=6).value = 'longitude'
# wb.save('example.xlsx')

wb=Workbook()  #创建一个工作表
ws=wb.active   #ws操作sheet页

#使用 Workbook.create_sheet() 创建新的worksheets
ws1=wb.create_sheet('Mysheetfirst',0)
ws2=wb.create_sheet('MysheetEnd')# insert at the end(default)
#修改sheet名字 Worksheet.title
ws1.title='新名字'
#修改sheet页标签的背景颜色
ws1.sheet_properties.tabColor='FF0000'
#打印Excel的所有sheet页名字 Worksbook.sheetnames
print(wb.sheetnames)
#给定一个worksheet的名字，可以把sheet赋值
ws3=wb['Sheet']
ws3.title='默认Sheet页'
#循环sheet页
for sheet in wb:
print(sheet.title)
#复制sheet页
ws4=wb.copy_worksheet(ws1)
wb.save('测试用.xlsx')

dataclean_path = 'F:\\年鉴自动处理\\2年鉴数据清理'
wb1 = Workbook()
sheet = wb1.get_sheet_by_name(wb1.sheetnames[0])
print(sheet)


wb1.save(dataclean_path + '\\' + '汇总.xlsx')




