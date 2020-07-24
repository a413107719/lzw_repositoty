import arcpy
import os
import openpyxl
import time
import socket
import uuid
import shutil
import gc


# 新建一个output数据库
def create_gdb(cad_folder, gdbpath):
    if arcpy.Exists("ProjectGDB.gdb"):
        shutil.rmtree(gdbpath)
        print('ProjectGDB.gdb数据库已存在,将删除')
    arcpy.CreateFileGDB_management(cad_folder, "ProjectGDB.gdb")
    print("成功创建ProjectGDB.gdb数据库")


# 提取点、线、面和注记数据
def dwg2gdb(cad_folder, gdb_path, muban_excel):
    wb = openpyxl.load_workbook(muban_excel)

    # 提取cad列表
    os.chdir(cad_folder)
    cads = []
    for i in ["*.dwg", "*.dgn", "*.dxf"]:
        cads = cads + arcpy.ListFiles(i)
    print('文件夹中存在CAD数据：' + str(cads) + '\n')

    # 分点、线、面、注记提取数据
    feature_kinds = ['Point', 'Polyline', 'Polygon', 'Annotation']
    for feature_kind in feature_kinds:
        num = len(cads)
        sheet = wb[feature_kind]
        max_row = sheet.max_row
        for cad in cads:
            num = num -1
            print('\n' + "-------------------------" + feature_kind + "还剩：" + str(num) + ' 个数据需要提取------------------------')
            print('【正在从"' + cad + '"中提取' + feature_kind + '数据】')
            oldtime = time.time()
            dwg_feature = cad_folder + cad + "\\" + feature_kind
            arcpy.env.workspace = gdb_path
            arcpy.env.overwriteOutput = True
            for name in range(2, max_row+1):
                time2 = time.time()
                sheetvalue = sheet.cell(name, 1).value
                if sheetvalue is None:
                    continue
                cadname = cad.split('.')[0]
                feature_name = sheetvalue + '_' + cadname
                selet_condition = sheet.cell(name, 2).value
                if feature_name is None:
                    pass
                else:
                    # 去除数量为空的数据
                    time31 = time.time()
                    print('Excel执行时间：%f' % (time31 - time2))
                    selection = arcpy.SelectLayerByAttribute_management(dwg_feature, "NEW_SELECTION", selet_condition)
                    time3 = time.time()
                    print('select执行时间：%f' % (time3 - time31))
                    rowcount1 = arcpy.GetCount_management(selection)
                    if int(str(rowcount1)) > 0:   # 因为rowcount1返回的是一个result格式
                        arcpy.FeatureClassToFeatureClass_conversion(dwg_feature, gdb_path, feature_name, selet_condition)
                        print("成功在数据库中添加要素：" + feature_name + '  数量：' + str(rowcount1))
                del name
                gc.collect()
            print('总执行时间：%f' % (time.time()-oldtime))
            del selection
            gc.collect()
            arcpy.Delete_management('in_memory')
        del cad
        gc.collect()
        print('-----------------------' + '\n' + '【' + feature_kind + "数据已提取完成" + '】' + "\n")

    # 删除空数据
    print('【' + "正在删除空数据" + '】')
    arcpy.env.workspace = gdb_path
    for shpfile in arcpy.ListFeatureClasses():
        rowcount = arcpy.GetCount_management(shpfile)
        # print("数据名称： " + shpfile + "     数据量：" + str(rowcount))
        if int(str(rowcount)) == 0:
            fullpath = gdb_path + '\\' + shpfile
            arcpy.Delete_management(fullpath)  # 在GDB中删除数据
            print(shpfile + "： 由于数据为空，已在GDB中删除")
    print("--------------------------CAD数据入库已完成------------------------", '\n')


# 清理字段
def clean_data(gdbfeatures):
    print('-----------------------开始清理数据--------------------------------')
    wb = openpyxl.load_workbook(muban_excel)
    feature_kinds = ['Point', 'Polyline', 'Polygon', 'Annotation']
    for feature_kind in feature_kinds:
        sheet = wb[feature_kind]
        max_row = sheet.max_row
        for row_num in range(2, max_row + 1):
            row_name = sheet.cell(row_num, 1).value
            contain_field = sheet.cell(row_num, 3).value  # 保留字段

            for num, featurename in enumerate(gdbfeatures):
                if featurename.split('_')[0] == row_name:
                    print('[' + '开始清理： "' + featurename + ']')
                    # 清理数据字段
                    if contain_field == ' ' or contain_field is None:
                        print('数据字段未清理')
                    else:
                        remain_field_list = contain_field.split(',') + ['OBJECTID', 'Shape', 'Shape_Length',
                                                                        'Shape_area', 'Shape_Area']
                        print("需要保留的字段包括：" + str(remain_field_list))
                        # 生成字段名列表
                        original_field_list = [f.name for f in arcpy.ListFields(featurename)]
                        print('数据原始字段包括：' + str(original_field_list))
                        for i in remain_field_list:
                            if i in original_field_list:
                                original_field_list.remove(i)
                        delete_field_list = original_field_list
                        if len(delete_field_list) != 0:
                            print('删除字段包括：' + str(delete_field_list))
                            arcpy.DeleteField_management(featurename, delete_field_list)
                        print('数据字段已清理', '\n')
    print("--------------------------数据清理完成------------------------", '\n')


# 分类合并数据
def merge_kind_features(gdbfeatures, gdb_path, cad_path):
    global shp_path
    print('-----------------------开始合并数据-------------------------')
    # 分类数据
    dic = {}
    for num, name in enumerate(gdbfeatures):
        kindname = name.split('_')[0]
        if kindname not in dic.keys():
            dic[kindname] = []
        dic[kindname].append(name)
    for i in dic.keys():
        print(i, dic[i])
    print('分类完毕。。。。', '\n', '\n')

    # 合并数据
    shp_path = cad_path + 'shp' + '\\'
    if os.path.isdir(shp_path):
        shutil.rmtree(shp_path)
    os.makedirs(shp_path)
    for key in dic.keys():
        dic[key].reverse()
        arcpy.env.workspace = gdb_path
        arcpy.Merge_management(dic[key], shp_path + key + '.shp')
        print("Finish Merge：", key, dic[key])
    print("--------------------------合并数据已完成------------------------", '\n')


# 新建map + 添加数据 + 可视化
def data_visualization(cad_folder, gdb_path, muban_excel, muban_aprx):
    global shp_path
    pro_aprx = arcpy.mp.ArcGISProject(muban_aprx)

    # 在map中添加数据
    print('【' + '开始在map中添加数据' + '】')
    data_map = pro_aprx.listMaps("Map")[0]
    arcpy.env.workspace = shp_path
    for feature_name in arcpy.ListFeatureClasses():
        feature_location = shp_path + feature_name
        print("shplocation" + feature_location)
        data_map.addDataFromPath(feature_location)
        print("成功添加数据: " + feature_name)
    print('【' + "数据添加完成" + '】' + '\n')
    data_layers = data_map.listLayers()

    # 设置数据样式
    print('【' + '开始设置符号样式' + '】')
    wb = openpyxl.load_workbook(muban_excel)
    os.chdir(cad_folder)
    feature_kinds = ['Point', 'Polyline', 'Polygon', 'Annotation']
    for feature_kind in feature_kinds:
        sheet = wb[feature_kind]
        max_row = sheet.max_row
        for row_num in range(2, max_row + 1):
            row_name = sheet.cell(row_num, 1).value
            for layer in data_layers:
                layer_name = str(layer)
                if layer_name == row_name:  # 根据excel表在GDB中选择相应的数据做符号化和标注设置
                    print('[' + '开始设置 "' + layer_name + '" 的符号和标注样式' + ']')
                    sheet_lable_size = sheet.cell(row_num, 4).value  # 设置size
                    sheet_color_way = sheet.cell(row_num, 5).value  # 渲染方式
                    sheet_color_field = sheet.cell(row_num, 6).value  # 渲染字段
                    sheet_color_ramp = sheet.cell(row_num, 7).value  # 色带和符号库样式
                    lable_ornot = sheet.cell(row_num, 8).value  # 标注与否
                    lable_field = sheet.cell(row_num, 9).value  # 标注字段

                    # 设置符号
                    sym = layer.symbology
                    if sheet_color_way == 'GraduatedColorsRenderer':
                        sym.updateRenderer(sheet_color_way)
                        sym.renderer.classificationField = sheet_color_field
                        sym.renderer.breakCount = 20
                        sym.renderer.colorRamp = pro_aprx.listColorRamps(sheet_color_ramp)[0]
                    elif sheet_color_way == 'UniqueValueRenderer':
                        sym.updateRenderer(sheet_color_way)
                        sym.renderer.fields = [sheet_color_field]
                        sym.renderer.colorRamp = pro_aprx.listColorRamps(sheet_color_ramp)[0]
                    else:
                        sym.renderer.symbol.applySymbolFromGallery(sheet_color_ramp)
                        # 获取并设置点符号大小、线符号宽度或面符号的轮廓宽度。
                        sym.renderer.symbol.size = sheet_lable_size
                    layer.symbology = sym
                    print('数据符号已设置')

                    # 设置标注
                    if layer.supports("SHOWLABELS"):
                        if lable_ornot == '是':
                            lblclass = layer.listLabelClasses()[0]
                            lblclass.expression = '$feature.' + str(lable_field)
                            lblclass.visible = True
                            layer.showLabels = True
                            print("图层标注已打开" + '\n')
                        else:
                            print("图层标注未设置" + '\n')
    print('【' + '符号样式设置已完成' + '】' + '\n')

    # 另存为新的项目文件
    time_text = time.strftime('%Y%m%d_%H%M', time.localtime(time.time()))
    file_name = "地形图信息提取" + time_text + ".aprx"
    pro_aprx.saveACopy(cad_folder + file_name)
    print('地形图信息提取完成，并已另存为aprx文件：' + file_name)
    del pro_aprx


def main_function():
    gdb_path = cadfolder + 'ProjectGDB.gdb'
    arcpy.env.workspace = cadfolder
    create_gdb(cadfolder, gdb_path)  # 新建一个gdbpath数据库
    dwg2gdb(cadfolder, gdb_path, muban_excel)  # 找到所有dwg文件，并提取信息到数据库
    arcpy.env.workspace = gdb_path
    gdb_features = arcpy.ListFeatureClasses()
    print(gdb_features)
    clean_data(gdb_features)
    merge_kind_features(gdb_features, gdb_path, cadfolder)
    data_visualization(cadfolder, gdb_path, muban_excel, muban_aprx)   # 新建aprx工程文件，新建map，数据可视化


if __name__ == '__main__':
    cadfolder = 'F:\\测试数据\\地形图信息提取\\cad\\'
    muban_excel = 'F:\\测试数据\\地形图信息提取\\模板\\地形图提取模板.xlsx'
    muban_aprx = 'F:\\测试数据\\地形图信息提取\\模板\\模板.aprx'
    main_function()




