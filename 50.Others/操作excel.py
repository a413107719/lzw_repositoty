import openpyxl,os


os.chdir('C:\\Users\\lzw19\\Desktop')
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']         # 选择一个excel页面
# 填入数据
max_row = sheet.max_row
newrow = max_row + 1
sheet.cell(row=newrow, column=1).value = 'ID'
sheet.cell(row=newrow, column=2).value = 'name.string'
sheet.cell(row=newrow, column=3).value = 'province'
sheet.cell(row=newrow, column=4).value = 'city'
sheet.cell(row=newrow, column=5).value = 'lattitude'
sheet.cell(row=newrow, column=6).value = 'longitude'
wb.save('example.xlsx')